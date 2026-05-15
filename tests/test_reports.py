from pathlib import Path

from openpyxl import load_workbook

from naukri_assistant.models import JobCandidate
from naukri_assistant.reports import ExternalJobReportRow, ReportExporter
from naukri_assistant.storage import LocalStore


def _candidate() -> JobCandidate:
    return JobCandidate(
        stable_key="job:123",
        job_id="123",
        url="https://www.naukri.com/job-listings-software-engineer-acme-123",
        title="Software Engineer",
        company="Acme Technologies",
        snippet="Python FastAPI AWS",
        raw_text="Python FastAPI AWS 1-3 years",
        experience_text="1-3 years",
    )


def test_external_workbook_exports_and_deduplicates(tmp_path: Path) -> None:
    store = LocalStore(tmp_path)
    exporter = ReportExporter(store)
    row = ExternalJobReportRow.from_candidate(
        _candidate(),
        apply_link="https://careers.example.com/jobs/123",
    )

    first = exporter.export_external_jobs([row])
    second = exporter.export_external_jobs([row])

    assert first.appended == 1
    assert first.duplicates == 0
    assert second.appended == 0
    assert second.duplicates == 1

    workbook = load_workbook(store.external_jobs_report_path)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    assert values[0] == (
        "Job name",
        "Company name",
        "Job ID",
        "Apply link",
        "Experience required",
        "Tech stack/skills",
        "Source",
        "Status",
    )
    assert values[1][0] == "Software Engineer"
    assert values[1][2] == "123"
    assert values[1][3] == "https://careers.example.com/jobs/123"
    assert values[1][7] == "External Review Needed"


def test_run_report_exports_latest_and_timestamped_files(tmp_path: Path) -> None:
    store = LocalStore(tmp_path)
    exporter = ReportExporter(store)
    report_path = exporter.export_run_report({"summary": {"processed": 3}})

    assert report_path.exists()
    assert store.latest_run_report_path.exists()
    assert '"processed": 3' in report_path.read_text(encoding="utf-8")
