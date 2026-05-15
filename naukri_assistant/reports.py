from __future__ import annotations

import json
import logging
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .logging_utils import LOGGER_NAME
from .models import JobCandidate
from .resume import KNOWN_SKILLS
from .storage import LocalStore
from .text_utils import canonical_url, normalize_text, stable_hash


logger = logging.getLogger(LOGGER_NAME)

EXTERNAL_HEADERS = [
    "Job name",
    "Company name",
    "Job ID",
    "Apply link",
    "Experience required",
    "Tech stack/skills",
    "Source",
    "Status",
]


@dataclass(slots=True)
class ExternalJobReportRow:
    job_name: str
    company_name: str
    job_id: str
    apply_link: str
    experience_required: str
    tech_stack_skills: str
    source: str = "External-company apply link"
    status: str = "External Review Needed"

    @classmethod
    def from_candidate(cls, candidate: JobCandidate, *, apply_link: str) -> "ExternalJobReportRow":
        return cls(
            job_name=candidate.title,
            company_name=candidate.company or "Unknown",
            job_id=candidate.job_id or "",
            apply_link=apply_link or candidate.url,
            experience_required=candidate.experience_text or "",
            tech_stack_skills=", ".join(extract_tech_stack(candidate)),
        )

    def dedupe_key(self) -> str:
        if self.job_id:
            return f"job:{self.job_id}"
        if self.apply_link:
            return f"link:{canonical_url(self.apply_link)}"
        fallback = "|".join([normalize_text(self.job_name), normalize_text(self.company_name)])
        return f"fallback:{stable_hash(fallback)}"

    def to_excel_row(self) -> list[str]:
        return [
            self.job_name,
            self.company_name,
            self.job_id,
            self.apply_link,
            self.experience_required,
            self.tech_stack_skills,
            self.source,
            self.status,
        ]


@dataclass(slots=True)
class ExcelExportResult:
    path: Path
    appended: int = 0
    duplicates: int = 0


class ReportExporter:
    def __init__(self, store: LocalStore) -> None:
        self.store = store
        self.store.ensure_layout()

    def export_external_jobs(self, rows: Iterable[ExternalJobReportRow]) -> ExcelExportResult:
        result = ExcelExportResult(path=self.store.external_jobs_report_path)
        rows = list(rows)
        workbook, worksheet = self._load_or_create_external_workbook()
        existing_keys = self._existing_external_keys(worksheet)

        for row in rows:
            dedupe_key = row.dedupe_key()
            if dedupe_key in existing_keys:
                result.duplicates += 1
                logger.info("Skipped duplicate external export row dedupe_key=%s", dedupe_key)
                continue
            worksheet.append(row.to_excel_row())
            existing_keys.add(dedupe_key)
            result.appended += 1

        self._style_external_sheet(worksheet)
        workbook.save(result.path)
        logger.info(
            "External workbook updated path=%s appended=%s duplicates=%s",
            result.path,
            result.appended,
            result.duplicates,
        )
        return result

    def export_run_report(self, payload: dict) -> Path:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        report_path = self.store.reports_dir / f"run-report-{timestamp}.json"
        serialized = json.dumps(payload, indent=2, ensure_ascii=True)
        report_path.write_text(serialized, encoding="utf-8")
        self.store.latest_run_report_path.write_text(serialized, encoding="utf-8")
        logger.info("Run report exported path=%s", report_path)
        return report_path

    def _load_or_create_external_workbook(self):
        if self.store.external_jobs_report_path.exists():
            workbook = load_workbook(self.store.external_jobs_report_path)
            worksheet = workbook.active
            return workbook, worksheet

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "External Review Needed"
        worksheet.append(EXTERNAL_HEADERS)
        return workbook, worksheet

    @staticmethod
    def _existing_external_keys(worksheet) -> set[str]:
        keys: set[str] = set()
        if worksheet.max_row < 2:
            return keys
        headers = [cell.value for cell in worksheet[1]]
        index = {header: position for position, header in enumerate(headers)}
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            job_id = str(values[index["Job ID"]] or "").strip() if "Job ID" in index else ""
            apply_link = str(values[index["Apply link"]] or "").strip() if "Apply link" in index else ""
            job_name = str(values[index["Job name"]] or "").strip() if "Job name" in index else ""
            company_name = str(values[index["Company name"]] or "").strip() if "Company name" in index else ""
            existing = ExternalJobReportRow(
                job_name=job_name,
                company_name=company_name,
                job_id=job_id,
                apply_link=apply_link,
                experience_required="",
                tech_stack_skills="",
            )
            keys.add(existing.dedupe_key())
        return keys

    @staticmethod
    def _style_external_sheet(worksheet) -> None:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
        widths = {
            "A": 32,
            "B": 28,
            "C": 18,
            "D": 48,
            "E": 22,
            "F": 38,
            "G": 28,
            "H": 26,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width


def extract_tech_stack(candidate: JobCandidate) -> list[str]:
    blob = normalize_text(" ".join([candidate.title, candidate.snippet, candidate.raw_text]))
    return sorted({skill for skill in KNOWN_SKILLS if normalize_text(skill) in blob})
